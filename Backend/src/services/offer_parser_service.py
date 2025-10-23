"""
Offer Parser Service - Adaptive Multi-Format Parser
Handles Imperial (hierarchical), Beautik (flat), and CCC (detailed) formats
Now handles both CSV and single-sheet Excel for flat/hierarchical data.
"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.offer_models import (
    ParsedOffer,
    ProjectMetadata,
    CostBreakdown,
    CostCategory,
    CostItem,
    DetailLevel,
    CategoryType
)
from src.utils.currency_utils import parse_eur_value, sum_eur_values
from src.services.classification_engine import (
    classify_item,
    normalize_unit,
    get_category_name
)


# ============================================================================
# FORMAT DETECTION
# ============================================================================

class OfferFormat:
    """Detected offer format type"""
    HIERARCHICAL_CSV = "hierarchical_csv"  # Imperial: A/B categories with items
    FLAT_CSV = "flat_csv"                  # Beautik: No categories, flat list
    DETAILED_EXCEL = "detailed_excel"      # CCC: Multi-sheet with unit prices
    SINGLE_SHEET_EXCEL = "single_sheet_excel" # NEW: Adapter for flat/hierarchical Excel


class FormatDetector:
    """Detects offer format from file content"""
    
    @staticmethod
    def _detect_csv_structure(file_content: bytes) -> str:
        """Helper to check if CSV is flat or hierarchical"""
        try:
            content_str = file_content.decode('utf-8')
        except UnicodeDecodeError:
            content_str = file_content.decode('latin-1')
        
        reader = csv.reader(io.StringIO(content_str))
        rows = list(reader)
        
        # Look for hierarchical markers (A, B, C categories)
        has_categories = False
        for row in rows[:30]:  # Check first 30 rows
            if len(row) > 1:
                cell_b = row[1].strip() if len(row) > 1 else ""
                if cell_b in ['A', 'B', 'C']:
                    has_categories = True
                    break
        
        if has_categories:
            return OfferFormat.HIERARCHICAL_CSV
        else:
            return OfferFormat.FLAT_CSV
            
    @staticmethod
    def detect(file_content: bytes, filename: str) -> str:
        """
        Detect offer format
        
        Returns:
            OfferFormat constant
        """
        filename_lower = filename.lower()
        is_excel = filename_lower.endswith(('.xlsx', '.xls'))
        is_csv = filename_lower.endswith('.csv')
        
        if is_excel:
            try:
                # We use read_only=True for fast detection
                wb = load_workbook(io.BytesIO(file_content), read_only=True)
                # Check for detailed multi-sheet format first
                if len(wb.sheetnames) > 2:
                    # Multi-sheet workbook = detailed format (CCC)
                    return OfferFormat.DETAILED_EXCEL
                else:
                    # Simple single-sheet Excel (Sinsay, or Imperial-in-Excel)
                    # Let the new adapter parser handle it
                    return OfferFormat.SINGLE_SHEET_EXCEL
            except Exception:
                # Corrupt Excel file, let adapter parser handle failure
                return OfferFormat.SINGLE_SHEET_EXCEL
        
        if is_csv:
            try:
                # It's a CSV, check if it's flat or hierarchical
                return FormatDetector._detect_csv_structure(file_content)
            except Exception:
                # Fallback for unparseable CSV
                return OfferFormat.FLAT_CSV
        
        # Fallback for unknown extensions (e.g., .txt)
        try:
            return FormatDetector._detect_csv_structure(file_content)
        except Exception:
            # Not CSV-like text, default to flat
            return OfferFormat.FLAT_CSV


# ============================================================================
# BASE PARSER CLASS
# ============================================================================

class BaseParser:
    """Base class for all parsers"""
    
    def __init__(self):
        self.warnings = []
    
    def add_warning(self, message: str):
        """Add a warning message"""
        self.warnings.append(message)
    
    def parse(self, file_content: bytes, filename: str) -> ParsedOffer:
        """Parse offer file - must be implemented by subclasses"""
        raise NotImplementedError


# ============================================================================
# HIERARCHICAL CSV PARSER (Imperial Brands)
# ============================================================================

class HierarchicalParser(BaseParser):
    """Parses Imperial Brands format: A/B/C categories with nested items"""
    
    def parse(self, file_content: bytes, filename: str) -> ParsedOffer:
        """Parse hierarchical CSV format"""
        # Decode CSV
        try:
            content_str = file_content.decode('utf-8')
        except UnicodeDecodeError:
            content_str = file_content.decode('latin-1')
            self.add_warning("File was not valid UTF-8, used latin-1 fallback.")
            
        reader = csv.reader(io.StringIO(content_str))
        rows = list(reader)
        
        # Extract metadata (rows 1-7)
        metadata = self._extract_metadata(rows)
        
        # Find table start (look for SUMMARY marker)
        table_start = self._find_table_start(rows)
        
        # Parse cost data
        categories, grand_total = self._parse_cost_data(rows, table_start)
        
        # Create cost breakdown
        cost_breakdown = CostBreakdown(
            categories=categories,
            grand_total_eur=grand_total
        )
        
        # Generate offer ID
        offer_id = self._generate_offer_id()
        
        # Create parsed offer
        return ParsedOffer(
            offer_id=offer_id,
            project=metadata,
            cost_breakdown=cost_breakdown,
            parsed_at=datetime.now(),
            source_filename=filename,
            detail_level=DetailLevel.SUMMARY,
            warnings=self.warnings
        )
    
    def _extract_metadata(self, rows: List[List[str]]) -> ProjectMetadata:
        """Extract project metadata from header rows"""
        metadata = {}
        
        # Parse rows 1-7 looking for key-value pairs
        for i in range(min(10, len(rows))):
            if len(rows[i]) < 2:
                continue
            
            cell_b = rows[i][1].strip()
            
            # Look for pattern "Key: Value"
            if ':' in cell_b:
                parts = cell_b.split(':', 1)
                key = parts[0].strip().lower()
                value = parts[1].strip() if len(parts) > 1 else ""
                
                if 'project' in key:
                    metadata['project_name'] = value
                elif 'beneficary' in key or 'client' in key or 'catre' in key.lower():
                    metadata['client_name'] = value
                elif 'address' in key or 'locatie' in key.lower():
                    metadata['address'] = value
                elif 'designer' in key:
                    metadata['lead_designer'] = value
                elif 'management' in key:
                    metadata['project_manager'] = value
                elif 'investitia' in key.lower():
                    metadata['project_name'] = value
        
        # Ensure project_name exists
        if 'project_name' not in metadata:
            metadata['project_name'] = "Unknown Project"
            self.add_warning("Project name not found in metadata")
        
        return ProjectMetadata(**metadata)
    
    def _find_table_start(self, rows: List[List[str]]) -> int:
        """Find row where cost table starts"""
        for i, row in enumerate(rows):
            if len(row) > 1:
                cell_b = row[1].strip().upper()
                # Look for SUMMARY or CENTRALIZATOR marker
                if 'SUMMARY' in cell_b or 'CENTRALIZATOR' in cell_b:
                    # Table starts 2 rows after marker
                    return i + 2
        
        self.add_warning("Could not find table marker, assuming row 10")
        return 10
    
    def _parse_cost_data(
        self, 
        rows: List[List[str]], 
        start_row: int
    ) -> Tuple[List[CostCategory], float]:
        """Parse hierarchical cost structure"""
        categories = []
        current_category = None
        current_items = []
        grand_total = 0.0
        
        for i in range(start_row, len(rows)):
            row = rows[i]
            if len(row) < 4: # Need at least 4 columns (B, C, D)
                continue
            
            cell_b = row[1].strip() if len(row) > 1 else ""
            cell_c = row[2].strip() if len(row) > 2 else ""
            cell_d = row[3].strip() if len(row) > 3 else ""
            
            # Check for grand total
            if 'TOTAL' in cell_b.upper() and 'w/o VAT' in cell_b:
                grand_total = parse_eur_value(cell_d) or 0.0
                break
            
            # Check for category header (single letter A, B, C)
            if cell_b in ['A', 'B', 'C']:
                # Save previous category
                if current_category and current_items:
                    category = CostCategory(
                        category_id=current_category['id'],
                        name=current_category['name'],
                        items=current_items,
                        total_eur=current_category['total']
                    )
                    categories.append(category)
                
                # Start new category
                category_total = parse_eur_value(cell_d) or 0.0
                current_category = {
                    'id': cell_b,
                    'name': cell_c,
                    'total': category_total
                }
                current_items = []
                continue
            
            # Check for item (numeric item number)
            if cell_b.isdigit() and current_category:
                value = parse_eur_value(cell_d)
                
                # Skip items with no value (intentional blanks)
                if value is None or value == 0:
                    continue
                
                # Classify item
                category_type, item_type = classify_item(cell_c, current_category['id'])
                
                item = CostItem(
                    item_number=cell_b,
                    description=cell_c,
                    value_eur=value,
                    category_id=current_category['id'],
                    item_type=item_type
                )
                current_items.append(item)
        
        # Save last category
        if current_category and current_items:
            category = CostCategory(
                category_id=current_category['id'],
                name=current_category['name'],
                items=current_items,
                total_eur=current_category['total']
            )
            categories.append(category)
        
        # Validate grand total
        if grand_total == 0:
            grand_total = sum(cat.total_eur for cat in categories)
            self.add_warning("Grand total not found, calculated from categories")
        
        return categories, grand_total
    
    def _generate_offer_id(self) -> str:
        """Generate unique offer ID: OFF_YYYYMMDD_NNN"""
        now = datetime.now()
        date_str = now.strftime("%Y%m%d")
        # TODO: Query Firestore for counter
        counter = 1
        return f"OFF_{date_str}_{counter:03d}"


# ============================================================================
# FLAT CSV PARSER (Beautik / Sinsay)
# ============================================================================

# ============================================================================
# FLAT CSV PARSER (Beautik / Sinsay)
# ============================================================================

# ============================================================================
# FLAT CSV PARSER (Beautik / Sinsay)
# ============================================================================

class FlatParser(BaseParser):
    """
    Parses flat list formats (Beautik/SL3, Sinsay).
    Auto-detects column positions based on headers.
    Can parse summary totals or full unit prices.
    """
    
    def __init__(self):
        super().__init__()
        # Column indices to find
        self.col_idx_item_num = -1
        self.col_idx_desc = -1
        self.col_idx_value = -1
        # New columns for unit pricing
        self.col_idx_unit = -1
        self.col_idx_qty = -1
        self.col_idx_unit_price = -1
        
        self.has_unit_pricing = False
        self.debug_logs = [] # NEW: For logging

    def parse(self, file_content: bytes, filename: str) -> ParsedOffer:
        """Parse flat CSV format"""
        try:
            content_str = file_content.decode('utf-8')
        except UnicodeDecodeError:
            # Fallback for files saved with non-standard encoding
            content_str = file_content.decode('latin-1')
            self.add_warning("File was not valid UTF-8, used latin-1 fallback.")
            
        reader = csv.reader(io.StringIO(content_str))
        rows = list(reader)
        
        # Extract metadata
        metadata = self._extract_metadata(rows)
        
        # Find table start and map columns
        table_start = self._find_table_start_and_map_cols(rows)
        
        # --- NEW DEBUGGING ---
        self.add_warning(f"DEBUG: Found table start at row {table_start}")
        mapped_cols = f"ItemNum={self.col_idx_item_num}, Desc={self.col_idx_desc}, Value={self.col_idx_value}, " \
                      f"Qty={self.col_idx_qty}, Unit={self.col_idx_unit}, UnitPrice={self.col_idx_unit_price}"
        self.add_warning(f"DEBUG: Mapped columns: {mapped_cols}")
        if table_start < len(rows):
             self.add_warning(f"DEBUG: First data row (raw): {rows[table_start]}")
        # --- END DEBUGGING ---
        
        # We must find these 3 columns to proceed
        if self.col_idx_item_num == -1 or self.col_idx_desc == -1 or self.col_idx_value == -1:
            self.add_warning("Could not find all required columns (Item#, Description, Value). Trying legacy format.")
            # Set default indices for legacy Beautik format as a fallback
            self.col_idx_item_num = 1  # Col B
            self.col_idx_desc = 2      # Col C
            self.col_idx_value = 3     # Col D
            
            # Check if these defaults are even valid
            max_cols = max(len(r) for r in rows) if rows else 0
            if max(self.col_idx_item_num, self.col_idx_desc, self.col_idx_value) >= max_cols:
                raise ValueError("Failed to parse flat CSV: Could not find required columns and legacy format is out of bounds.")
        
        # Check if we also found unit pricing columns
        if self.col_idx_qty != -1 and self.col_idx_unit_price != -1:
            self.has_unit_pricing = True
            
        # Parse items and auto-categorize
        categories, grand_total = self._parse_items(rows, table_start)
        
        cost_breakdown = CostBreakdown(
            categories=categories,
            grand_total_eur=grand_total
        )
        
        offer_id = self._generate_offer_id()
        
        # Set detail level based on what we found
        detail_level = DetailLevel.UNIT_PRICES if self.has_unit_pricing else DetailLevel.SUMMARY
        
        return ParsedOffer(
            offer_id=offer_id,
            project=metadata,
            cost_breakdown=cost_breakdown,
            parsed_at=datetime.now(),
            source_filename=filename,
            detail_level=detail_level,
            warnings=self.warnings
        )
    
    def _extract_metadata(self, rows: List[List[str]]) -> ProjectMetadata:
        """Extract metadata from Beautik format"""
        metadata = {}
        
        for i in range(min(15, len(rows))):
            if len(rows[i]) < 2:
                continue
            
            # Check for Beautik-style metadata (Key: Value in Col B/C)
            try:
                cell_b = rows[i][1].strip()
                cell_c = rows[i][2].strip() if len(rows[i]) > 2 else ""
            except IndexError:
                continue

            if ':' in cell_b:
                key = cell_b.replace(':', '').strip().lower()
                
                if 'catre' in key or 'client' in key:
                    metadata['client_name'] = cell_c
                elif 'investitia' in key or 'project' in key:
                    metadata['project_name'] = cell_c
                elif 'locatie' in key or 'location' in key:
                    metadata['address'] = cell_c
                elif 'data' in key or 'date' in key:
                    metadata['offer_date'] = cell_c
                elif 'oferta' in key or 'offer' in key:
                    metadata['offer_number'] = cell_c
        
        if 'project_name' not in metadata:
            metadata['project_name'] = metadata.get('client_name', "Unknown Project")
        
        return ProjectMetadata(**metadata)
    
    def _find_table_start_and_map_cols(self, rows: List[List[str]]) -> int:
        """
        Find table start row and map column indices from header.
        """
        # --- Define all known header variants ---
        header_keywords_item = ['nr.', 'lp', 'item', 'nr crt']
        header_keywords_desc = ['description', 'descriptions', 'denumire', 'descriere', 'detalii']
        header_keywords_value = ['total', 'value', 'valoare', 'val. totala', 'total value']
        
        # New keywords for unit pricing
        header_keywords_unit = ['j.m', 'unit', 'u.m', 'unit of measurement', 'um', 'unitate']
        header_keywords_qty = ['cantitate', 'qty', 'quantity', 'number of units', 'cant']
        header_keywords_unit_price = ['unit price', 'price/unit', 'val. unitara', 'pret unitar', 'pu']
        
        for i, row in enumerate(rows):
            if len(row) < 3:
                continue
            
            found_headers_count = 0
            
            for col_idx, cell in enumerate(row):
                cell_lower = cell.strip().lower().replace('"', '') # Clean cell
                
                if cell_lower in header_keywords_item:
                    self.col_idx_item_num = col_idx
                    found_headers_count += 1
                elif cell_lower in header_keywords_desc:
                    self.col_idx_desc = col_idx
                    found_headers_count += 1
                elif cell_lower in header_keywords_value:
                    self.col_idx_value = col_idx
                    found_headers_count += 1
                elif cell_lower in header_keywords_unit:
                    self.col_idx_unit = col_idx
                elif cell_lower in header_keywords_qty:
                    self.col_idx_qty = col_idx
                elif cell_lower in header_keywords_unit_price:
                    self.col_idx_unit_price = col_idx
            
            # We need at least 3 headers (item, desc, total) to call this a header row
            if found_headers_count >= 3:
                # We found a valid header row
                return i + 1  # Data starts on the next row
        
        # Fallback if no headers found
        self.add_warning("Could not detect header row. Assuming legacy format on row 12.")
        return 12
    
    def _is_item_number(self, s: str) -> bool:
        """Check if string is an item number (e.g. '1', '1.1', '8.2')"""
        s = s.strip()
        if not s:
            return False
        # Check if it's a digit or a float-like string (e.g., "1.1")
        return s.replace('.', '', 1).isdigit()

    def _safe_get_value(self, row: List[str], index: int) -> str:
        """Safely get value from row if index is valid"""
        if index != -1 and index < len(row):
            val = row[index]
            if val is None:
                return ""
            return str(val).strip()
        return ""
    
    def _parse_float(self, s: str) -> Optional[float]:
        """Parse a string into a float, handling commas and spaces"""
        if not s:
            return None
        try:
            # Clean currency symbols, letters, and all spaces
            s_cleaned = str(s)
            s_cleaned = s_cleaned.replace(' ', '')
            s_cleaned = s_cleaned.replace('€', '').replace('$', '')
            s_cleaned = s_cleaned.replace('PLN', '').replace('RON', '').replace('EUR', '')
            s_cleaned = s_cleaned.replace(',', '.')
        
            # Check if there are multiple '.' which means the first ones are
            # thousands separators (e.g., "1.234.567,89" -> "1.234.567.89")
            if s_cleaned.count('.') > 1:
                parts = s_cleaned.split('.')
                s_cleaned = "".join(parts[:-1]) + "." + parts[-1]
                
            return float(s_cleaned)
        except (ValueError, TypeError):
            # Try to handle cases like '8366.7069999999985' which might be read as text
            try:
                return float(s)
            except (ValueError, TypeError):
                # NEW DEBUGGING
                self.debug_logs.append(f"Failed to parse float: '{s}'")
                return None

    def _parse_items(
        self,
        rows: List[List[str]],
        start_row: int
    ) -> Tuple[List[CostCategory], float]:
        """Parse flat items and auto-categorize them using mapped columns"""
        items_by_category = {
            'A': [], 'B': [], 'C': []
        }
        grand_total = 0.0
        
        # Determine the maximum column index we'll need to read
        all_indices = [
            self.col_idx_item_num, self.col_idx_desc, self.col_idx_value,
            self.col_idx_unit, self.col_idx_qty, self.col_idx_unit_price
        ]
        max_idx_needed = max(idx for idx in all_indices if idx != -1)
        if max_idx_needed == -1: # Should be impossible due to check in parse()
            return [], 0.0 

        parsed_item_count = 0 # NEW DEBUGGING
        
        for i in range(start_row, len(rows)):
            row = rows[i]
            if len(row) <= max_idx_needed:
                continue
            
            try:
                item_num_str = self._safe_get_value(row, self.col_idx_item_num)
                desc_str = self._safe_get_value(row, self.col_idx_desc)
                total_value_str = self._safe_get_value(row, self.col_idx_value)
            except IndexError:
                continue # Row is shorter than expected

            # Check for total row (look in both item and desc columns)
            if 'total' in item_num_str.lower() or 'total' in desc_str.lower():
                grand_total = self._parse_float(total_value_str) or 0.0
                break
            
            # Check for item (is it a number like "1" or "1.1"?)
            if self._is_item_number(item_num_str):
                
                total_value = self._parse_float(total_value_str)
                
                if total_value is None or total_value == 0:
                    # Don't skip items with 0 value if they have unit pricing
                    if not (self.has_unit_pricing and total_value == 0):
                        # NEW DEBUGGING
                        if i < start_row + 10: # Log first few skips
                            self.debug_logs.append(f"Skipped row {i}: Item '{item_num_str}' has no value ('{total_value_str}')")
                        continue
                
                # --- Parse Unit Price Data (if columns exist) ---
                quantity_val = None
                unit_str = None
                unit_normalized = None
                unit_price_val = None

                if self.has_unit_pricing:
                    unit_str = self._safe_get_value(row, self.col_idx_unit)
                    quantity_str = self._safe_get_value(row, self.col_idx_qty)
                    unit_price_str = self._safe_get_value(row, self.col_idx_unit_price)

                    quantity_val = self._parse_float(quantity_str)
                    unit_price_val = self._parse_float(unit_price_str)
                    unit_normalized = normalize_unit(unit_str) if unit_str else None
                
                # Classify item to determine category
                category_type, item_type = classify_item(desc_str)
                
                item = CostItem(
                    item_number=item_num_str,
                    description=desc_str,
                    value_eur=total_value,
                    category_id=category_type.value,
                    item_type=item_type,
                    # Add unit price data
                    quantity=quantity_val,
                    unit=unit_str,
                    unit_normalized=unit_normalized,
                    unit_price_eur=unit_price_val
                )
                
                items_by_category[category_type.value].append(item)
                parsed_item_count += 1 # NEW DEBUGGING
        
        # Create categories from grouped items
        categories = []
        for cat_id in ['A', 'B', 'C']:
            items = items_by_category[cat_id]
            if items:
                total = sum(item.value_eur for item in items)
                category = CostCategory(
                    category_id=cat_id,
                    name=get_category_name(CategoryType(cat_id)),
                    items=items,
                    total_eur=total
                )
                categories.append(category)
        
        if grand_total == 0:
            grand_total = sum(cat.total_eur for cat in categories)
            
        # --- NEW DEBUGGING ---
        self.add_warning(f"DEBUG: Finished parsing. Parsed {parsed_item_count} items.")
        self.add_warning(f"DEBUG: Calculated Grand Total: {grand_total}")
        if self.debug_logs:
             self.add_warning("DEBUG Logs: " + " | ".join(self.debug_logs[:5])) # Show first 5 debug logs
        
        return categories, grand_total
    
    def _generate_offer_id(self) -> str:
        """Generate offer ID"""
        now = datetime.now()
        date_str = now.strftime("%Y%m%d")
        counter = 1
        return f"OFF_{date_str}_{counter:03d}"
    
# ============================================================================
# DETAILED EXCEL PARSER (CCC Pitesti)
# ============================================================================

class DetailedExcelParser(BaseParser):
    """Parses CCC format: Multi-sheet Excel with unit prices"""
    
    def parse(self, file_content: bytes, filename: str) -> ParsedOffer:
        """Parse detailed Excel format"""
        
        # *** FIX IS HERE: Added data_only=True ***
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        # Find summary sheet
        summary_sheet = self._find_summary_sheet(wb)
        
        # Extract metadata from summary
        metadata = self._extract_metadata(summary_sheet)
        
        # Parse summary totals
        summary_totals = self._parse_summary_totals(summary_sheet)
        
        # Parse detailed sheets
        categories = self._parse_detailed_sheets(wb, summary_totals)
        
        grand_total = sum(cat.total_eur for cat in categories)
        
        cost_breakdown = CostBreakdown(
            categories=categories,
            grand_total_eur=grand_total
        )
        
        offer_id = self._generate_offer_id()
        
        return ParsedOffer(
            offer_id=offer_id,
            project=metadata,
            cost_breakdown=cost_breakdown,
            parsed_at=datetime.now(),
            source_filename=filename,
            detail_level=DetailLevel.UNIT_PRICES,
            warnings=self.warnings
        )
    
    def _find_summary_sheet(self, wb) -> object:
        """Find the summary sheet"""
        for name in wb.sheetnames:
            name_upper = name.upper().strip()
            # ADDED 'CENTRALIZATOR'
            if 'SUMMARY' in name_upper or 'PODSUMOWANIE' in name_upper or 'CENTRALIZATOR' in name_upper:
                return wb[name]
        
        self.add_warning("Could not find 'SUMMARY' or 'CENTRALIZATOR' sheet, using first sheet.")
        # Fallback to first sheet
        return wb[wb.sheetnames[0]]
    
    def _extract_metadata(self, sheet) -> ProjectMetadata:
        """Extract metadata from summary sheet"""
        metadata = {}
        
        # Scan first 15 rows
        for row_idx in range(1, 16):
            for col_idx in range(1, 6):
                cell_value = sheet.cell(row_idx, col_idx).value
                if not cell_value:
                    continue
                
                cell_str = str(cell_value).strip()
                
                # Look for location/project info
                if 'location' in cell_str.lower() or 'lokalizacja' in cell_str.lower():
                    next_cell = sheet.cell(row_idx, col_idx + 1).value
                    if next_cell:
                        metadata['project_name'] = str(next_cell).strip()
                
                # Look for contractor
                if 'wykonawca' in cell_str.lower() or 'contractor' in cell_str.lower():
                    next_cell = sheet.cell(row_idx, col_idx + 1).value
                    if next_cell:
                        metadata['client_name'] = str(next_cell).strip()
        
        if 'project_name' not in metadata:
            metadata['project_name'] = "Unknown Project"
        
        return ProjectMetadata(**metadata)
    
    def _parse_summary_totals(self, sheet) -> Dict[str, float]:
        """Parse category totals from summary sheet"""
        totals = {}
        
        # Check header to see if this is Vivo format
        header_cell_b = str(sheet.cell(1, 2).value).strip().lower()
        is_vivo_format = 'categorii' in header_cell_b

        if is_vivo_format:
            # New VIVO format: Nr. | Categorii | Valoare
            for row_idx in range(2, 30): # Start from row 2
                cell_b_val = sheet.cell(row_idx, 2).value # Col B: Category Name
                if not cell_b_val:
                    continue # Stop when we run out of categories
                
                cell_c_val = sheet.cell(row_idx, 3).value # Col C: Value
                
                # Clean name to use as a key
                name_key = str(cell_b_val).lower().strip()
                value = float(cell_c_val) if cell_c_val else 0.0
                
                totals[name_key] = value
        
        else:
            # Old CCC format
            for row_idx in range(1, 30):
                cell_a = sheet.cell(row_idx, 1).value
                if not cell_a:
                    continue
                
                cell_str = str(cell_a).lower()
                cell_d = sheet.cell(row_idx, 4).value
                
                if 'budowlanka' in cell_str or 'construction' in cell_str:
                    totals['construction'] = float(cell_d) if cell_d else 0.0
                elif 'elektryka' in cell_str or 'electric' in cell_str:
                    totals['electrical'] = float(cell_d) if cell_d else 0.0
                elif 'dodatkowe' in cell_str or 'additional' in cell_str:
                    totals['additional'] = float(cell_d) if cell_d else 0.0
        
        return totals
    
    def _parse_detailed_sheets(self, wb, summary_totals: Dict) -> List[CostCategory]:
        """Parse detailed breakdown sheets based on summary totals"""
        categories = []
        
        # Loop through the categories found in the summary
        for category_key, category_total in summary_totals.items():
            
            sheet_keywords = []
            category_id = CategoryType.UNKNOWN.value # Default
            category_name = category_key.title()

            # Map summary keys to sheet names and Category IDs
            if 'arhitectura' in category_key:
                sheet_keywords = ['ARHITECTURA']
                category_id = CategoryType.A_ARCHITECTURAL.value
                category_name = 'Lucrari de Arhitectura'
            
            elif 'demolare' in category_key:
                sheet_keywords = ['DEMOLARE']
                category_id = CategoryType.A_ARCHITECTURAL.value # Group demo under Archi
                category_name = 'Lucrari de Demolare'
            
            elif 'mep' in category_key or 'instalatii' in category_key:
                sheet_keywords = ['MEP', 'INSTALATII']
                category_id = CategoryType.B_MEP.value
                category_name = 'Lucrari MEP'
            
            elif 'preliminarii' in category_key:
                sheet_keywords = ['PRELIMINARII']
                category_id = CategoryType.C_PROFESSIONAL.value
                category_name = 'Preliminarii'

            elif 'construction' in category_key or 'budowlanka' in category_key:
                # Handle old CCC format
                sheet_keywords = ['BUDOWLANKA', 'CONSTRUCTION']
                category_id = CategoryType.A_ARCHITECTURAL.value
                category_name = 'Construction Works'
                
            elif 'electric' in category_key or 'elektryka' in category_key:
                # Handle old CCC format
                sheet_keywords = ['ELEKTRYKA', 'ELECTRIC']
                category_id = CategoryType.B_MEP.value
                category_name = 'Electrical Works'

            if not sheet_keywords:
                self.add_warning(f"Skipping unknown category from summary: '{category_key}'")
                continue

            # Find the corresponding sheet
            sheet = self._find_sheet_by_keywords(wb, sheet_keywords)
            
            if sheet:
                # Find header row and format type
                header_row, format_type = self._find_header_row_and_format(sheet)
                if not header_row:
                    self.add_warning(f"Could not find header row in sheet: '{sheet.title}'")
                    continue
                
                # Parse items using the detected format
                items = self._parse_detail_sheet(sheet, header_row, format_type)
                
                if items:
                    # Use total from summary
                    category = CostCategory(
                        category_id=category_id,
                        name=category_name,
                        items=items,
                        total_eur=category_total 
                    )
                    categories.append(category)
            else:
                 self.add_warning(f"Could not find detail sheet for category: '{category_key}'")

        return categories
    
    def _find_sheet_by_keywords(self, wb, keywords: List[str]):
        """Find sheet by keywords"""
        for sheet_name in wb.sheetnames:
            name_upper = sheet_name.upper()
            for keyword in keywords:
                if keyword.upper() in name_upper:
                    return wb[sheet_name]
        return None
    
    def _parse_detail_sheet(self, sheet, header_row: int, format_type: str) -> List[CostItem]:
        """Parse detailed sheet with unit prices"""
        items = []
        
        # --- Define Column Mappings ---
        # (Col A is Item Num for both)
        if format_type == 'vivo':
            # Vivo format: Descriere(B), U.M(D), Cantitate(E), Pret Unitar(F), Valoare(G)
            col_desc = 2
            col_unit = 4
            col_qty = 5
            col_unit_price = 6
            col_total = 7
        else: # 'ccc' or default
            # CCC format: OPIS(C), Unit(D), Price(E), Qty(F), Total(H)
            col_desc = 3
            col_unit = 4
            col_qty = 6       # Note: Qty/Price are swapped vs Vivo
            col_unit_price = 5
            col_total = 8

        # Parse data rows
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            # Column A: Item number (LP)
            item_num = sheet.cell(row_idx, 1).value
            if not item_num:
                continue
            
            # --- Use mapped columns ---
            desc = sheet.cell(row_idx, col_desc).value
            if not desc:
                continue
            
            desc_str = str(desc).strip()
            
            unit = sheet.cell(row_idx, col_unit).value
            unit_str = str(unit).strip() if unit else None
            
            unit_price = sheet.cell(row_idx, col_unit_price).value
            unit_price_val = float(unit_price) if unit_price else None
            
            quantity = sheet.cell(row_idx, col_qty).value
            quantity_val = float(quantity) if quantity else None
            
            total_val = sheet.cell(row_idx, col_total).value
            total = float(total_val) if total_val else 0.0
            
            # Skip rows with no value
            if total == 0:
                continue
            
            # Classify item
            category_type, item_type = classify_item(desc_str)
            
            # Normalize unit
            unit_normalized = normalize_unit(unit_str) if unit_str else None
            
            item = CostItem(
                item_number=str(item_num),
                description=desc_str,
                value_eur=total,
                category_id=category_type.value,
                item_type=item_type,
                quantity=quantity_val,
                unit=unit_str,
                unit_normalized=unit_normalized,
                unit_price_eur=unit_price_val
            )
            
            items.append(item)
        
        return items
    
    def _find_header_row_and_format(self, sheet) -> Tuple[Optional[int], Optional[str]]:
        """
        Find header row in sheet and return its index and format type.
        """
        # Look for headers in first 10 rows
        for row_idx in range(1, 10):
            # Check for Vivo format (Descriere)
            # Scan columns B and C for "Descriere"
            for col_idx in [2, 3]: 
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if cell_str == 'descriere':
                        return row_idx, 'vivo'
            
            # Check for CCC format (OPIS / DESCRIPTION)
            cell_value_c = sheet.cell(row_idx, 3).value
            if cell_value_c:
                cell_str_c = str(cell_value_c).upper()
                if 'DESCRIPTION' in cell_str_c or 'OPIS' in cell_str_c:
                    return row_idx, 'ccc'

        return None, None
    
    def _generate_offer_id(self) -> str:
        """Generate offer ID"""
        now = datetime.now()
        date_str = now.strftime("%Y%m%d")
        counter = 1
        return f"OFF_{date_str}_{counter:03d}"

# ============================================================================
# NEW: SINGLE-SHEET EXCEL "BRIDGE" PARSER
# ============================================================================

class SingleSheetExcelParser(BaseParser):
    """
    Adapter parser for single-sheet Excel files.
    Converts the first sheet to CSV bytes in-memory, then routes to
    the appropriate CSV parser (Flat or Hierarchical).
    """

    def _convert_sheet_to_csv_bytes(self, sheet: Worksheet) -> bytes:
        """Converts an openpyxl sheet to in-memory CSV bytes"""
        csv_data = io.StringIO()
        writer = csv.writer(csv_data)
        
        for row in sheet.iter_rows():
            # Write cell values, handling None
            writer.writerow([cell.value if cell.value is not None else "" for cell in row])
            
        return csv_data.getvalue().encode('utf-8')

    def parse(self, file_content: bytes, filename: str) -> ParsedOffer:
        """
        Loads Excel, converts to CSV, and routes to the correct parser.
        """
        try:
            # *** FIX IS HERE: Added data_only=True ***
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
            
            first_sheet = wb[wb.sheetnames[0]]
            self.add_warning(f"Parsing single-sheet Excel, using sheet: '{first_sheet.title}'")
            
            # Convert the sheet to CSV bytes
            csv_bytes = self._convert_sheet_to_csv_bytes(first_sheet)
            
            # Now, detect the structure of the *CSV data*
            csv_format = FormatDetector._detect_csv_structure(csv_bytes)
            
            if csv_format == OfferFormat.HIERARCHICAL_CSV:
                self.add_warning("Detected Hierarchical structure in Excel file.")
                parser = HierarchicalParser()
            else:
                self.add_warning("Detected Flat structure in Excel file.")
                parser = FlatParser()
            
            # Call the chosen parser with the new CSV bytes
            parsed_offer = parser.parse(csv_bytes, filename)
            
            # Append warnings from this adapter and the sub-parser
            all_warnings = self.warnings + parser.warnings
            parsed_offer.warnings = all_warnings
            
            return parsed_offer

        except Exception as e:
            self.add_warning(f"Failed to parse single-sheet Excel: {e}")
            raise ValueError(f"Failed to read single-sheet Excel file: {str(e)}")


# ============================================================================
# MAIN PARSER SERVICE
# ============================================================================

class OfferParserService:
    """Main parser service - routes to appropriate parser"""
    
    def parse_offer(self, file_content: bytes, filename: str) -> ParsedOffer:
        """
        Parse offer file using adaptive format detection
        
        Args:
            file_content: Raw file bytes
            filename: Original filename
            
        Returns:
            ParsedOffer object
            
        Raises:
            ValueError: If parsing fails
        """
        # Detect format
        detected_format = FormatDetector.detect(file_content, filename)
        
        # Route to appropriate parser
        if detected_format == OfferFormat.HIERARCHICAL_CSV:
            parser = HierarchicalParser()
        elif detected_format == OfferFormat.FLAT_CSV:
            parser = FlatParser()
        elif detected_format == OfferFormat.DETAILED_EXCEL:
            parser = DetailedExcelParser()
        elif detected_format == OfferFormat.SINGLE_SHEET_EXCEL:
            # NEW: Route to the adapter
            parser = SingleSheetExcelParser()
        else:
            # Fallback to FlatParser for safety
            self.add_warning(f"Unknown format detected, falling back to FlatParser.")
            parser = FlatParser()
        
        # Parse
        try:
            parsed_offer = parser.parse(file_content, filename)
            
            # Validate
            errors = parsed_offer.validate()
            if errors:
                raise ValueError(f"Validation failed: {'; '.join(errors)}")
            
            return parsed_offer
            
        except Exception as e:
            # Re-raise with more context
            raise ValueError(f"Parsing failed for format '{detected_format}': {str(e)}")